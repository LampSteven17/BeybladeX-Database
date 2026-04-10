#!/usr/bin/env python3
"""
Import battle and stamina data from BEYBLADE-X_MASTER-LIST.xlsm into the
battles and stamina_trials tables.

The Excel aggregates results per matchup (e.g., "SPIN W: 10" means 10 spin
wins). This script expands each tally into individual battle rows so we can
query them with full flexibility (time-series, session analysis, etc.).

Usage:
    python scripts/import_battles_excel.py [path_to_xlsx]

Default path: /mnt/c/Users/RTX-MONSTER/Downloads/BEYBLADE-X_MASTER-LIST.xlsm
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from db import get_connection, init_schema

DEFAULT_PATH = "/mnt/c/Users/RTX-MONSTER/Downloads/BEYBLADE-X_MASTER-LIST.xlsm"

# Finish type → official BX point values
FINISH_POINTS = {"spin": 1, "over": 2, "xtreme": 3, "burst": 2}

# Column mapping for the BATTLES sheet.
# (col_index_1based, finish_type, result)
BATTLE_COLUMNS = [
    (8,  "spin",   "win"),    # SPIN W
    (10, "over",   "win"),    # OVER W
    (12, "xtreme", "win"),    # XTREME W
    (14, "burst",  "win"),    # BURST W
    (17, "spin",   "loss"),   # SPIN L
    (19, "over",   "loss"),   # OVER L
    (21, "xtreme", "loss"),   # EXTRME L
    (23, "burst",  "loss"),   # BURST L
]


def import_battles(conn, ws):
    """Import BATTLES sheet. Each tally cell becomes N individual rows."""
    total = 0
    for row_idx in range(2, ws.max_row + 1):
        combo_id = ws.cell(row=row_idx, column=1).value
        opponent_id = ws.cell(row=row_idx, column=3).value
        stadium = ws.cell(row=row_idx, column=5).value

        if not combo_id or not opponent_id:
            continue

        for col, finish_type, result in BATTLE_COLUMNS:
            count = ws.cell(row=row_idx, column=col).value
            if not count or not isinstance(count, (int, float)) or count <= 0:
                continue
            count = int(count)
            points = FINISH_POINTS[finish_type]
            for _ in range(count):
                conn.execute("""
                    INSERT INTO battles (combo_id, opponent_id, stadium,
                                        finish_type, result, points, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, 'excel-import')
                """, [combo_id, opponent_id, stadium, finish_type, result, points])
                total += 1

    return total


def parse_spin_time(time_str) -> int | None:
    """Parse spin time like '2:39:57' (M:SS:ms) to total milliseconds."""
    if not time_str or not isinstance(time_str, str):
        return None
    parts = time_str.strip().split(":")
    if len(parts) != 3:
        return None
    try:
        minutes = int(parts[0])
        seconds = int(parts[1])
        centis = int(parts[2])
        return (minutes * 60 + seconds) * 1000 + centis * 10
    except ValueError:
        return None


def import_stamina(conn, ws):
    """Import STAMINA sheet."""
    total = 0
    for row_idx in range(2, ws.max_row + 1):
        combo_id = ws.cell(row=row_idx, column=1).value
        if not combo_id:
            continue

        for trial_num, col in enumerate([3, 5, 7], start=1):
            time_val = ws.cell(row=row_idx, column=col).value
            if time_val is None:
                continue
            # Handle both string ("2:39:57") and datetime formats
            if hasattr(time_val, 'strftime'):
                # Excel sometimes parses times as datetime
                ms = (time_val.hour * 3600 + time_val.minute * 60 + time_val.second) * 1000
            else:
                ms = parse_spin_time(str(time_val))
            if ms is None or ms <= 0:
                continue
            conn.execute("""
                INSERT INTO stamina_trials (combo_id, spin_time_ms, trial_number)
                VALUES (?, ?, ?)
            """, [combo_id, ms, trial_num])
            total += 1

    return total


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    if not Path(path).exists():
        print(f"ERROR: File not found: {path}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    print(f"Loading {path}...")
    wb = openpyxl.load_workbook(path, data_only=True)

    conn = get_connection()
    init_schema(conn)

    # Clear previous imports to make this idempotent
    prev_battles = conn.execute("SELECT COUNT(*) FROM battles WHERE session_id = 'excel-import'").fetchone()[0]
    if prev_battles > 0:
        print(f"  Clearing {prev_battles} previously imported battles...")
        conn.execute("DELETE FROM battles WHERE session_id = 'excel-import'")
    prev_stamina = conn.execute("SELECT COUNT(*) FROM stamina_trials").fetchone()[0]
    if prev_stamina > 0:
        print(f"  Clearing {prev_stamina} previously imported stamina trials...")
        conn.execute("DELETE FROM stamina_trials")

    print("Importing BATTLES sheet...")
    battle_count = import_battles(conn, wb["BATTLES"])
    print(f"  {battle_count} individual battle rows created")

    print("Importing STAMINA sheet...")
    stamina_count = import_stamina(conn, wb["STAMINA"])
    print(f"  {stamina_count} stamina trials imported")

    conn.commit()
    conn.close()
    print(f"\nDone: {battle_count} battles + {stamina_count} stamina trials")


if __name__ == "__main__":
    main()
